"""Capacity-signal ingest for the data-center site search.

Five public datasets feed the capacity-aware ranking. None of them have a
stable single-URL API across years, so this module follows the same drop-in
pattern as tax_delinquent: the user downloads each year's bulk file (or
exports the public web tool to CSV) and drops it under
`data/capacity_inputs/<source>/`. This script then parses every file present
and emits normalized JSONLs the v3 enrich script consumes.

Sources & download instructions
================================

PJM interconnection queue (Phase 1.1)
  Drop directory: data/capacity_inputs/pjm_queue/*.xlsx (or .csv)
  Source:         https://www.pjm.com/planning/services-requests/interconnection-queues
                  - Click "Active Queue" -> Excel export
                  - Or use the Queue Scope geospatial tool (richer data)
  Required cols:  Queue Number, Project Name, MW (or MFO), POI Substation,
                  State (filter to VA), Status, Submitted, Withdraw Date
  Output:         data/pjm_queue.jsonl

EIA Form 860 generators (Phase 1.3)
  Drop directory: data/capacity_inputs/eia_860/*.xlsx
  Source:         https://www.eia.gov/electricity/data/eia860/
                  - Latest annual ZIP -> 3_1_Generator_Y2024.xlsx (or 2023)
  Required cols:  Plant Name, State, County, Nameplate Capacity (MW),
                  Status, Operating Year, Retirement Year, Latitude, Longitude
  Output:         data/eia_generators_va.jsonl

VADEQ water withdrawals (Phase 2)
  Drop directory: data/capacity_inputs/vadeq_water/*.csv
  Source:         https://www.deq.virginia.gov/water/water-quantity/water-withdrawal-reporting
                  - Annual CSV bulk download "VWUDS Annual Withdrawal Data"
  Required cols:  Facility Name, Source Type, MGY (millions of gallons/year)
                  or annual_withdrawal_mgd, Latitude, Longitude
  Output:         data/va_water_withdrawals.jsonl

VPDES discharge volumes (Phase 3)
  Drop directory: data/capacity_inputs/vpdes/*.csv
  Source:         https://www.deq.virginia.gov/permits-regulations/permits/water/vpdes
                  - "VPDES Permit Look-up" CSV export OR EPA ECHO
                    https://echo.epa.gov/files/echodownloads/echo_npdes_VA.zip
  Required cols:  Permit Number, Facility Name, Permitted Flow MGD,
                  Avg Discharge MGD (or Design Flow), Latitude, Longitude
  Output:         data/va_vpdes_volumes.jsonl

FCC broadband fiber (Phase 4)
  Drop directory: data/capacity_inputs/fcc_broadband/*.csv
  Source:         https://broadbandmap.fcc.gov/data-download/nationwide-data
                  - Pick latest version -> "Broadband Serviceable Locations"
                    CSV filtered to VA. Also: "Service Provider" CSV.
  Required cols:  Provider, Technology Code (filter 50/71/72 = fiber/wireline),
                  Latitude, Longitude
  Output:         data/va_fcc_fiber.jsonl

Run:
    python scripts/fetch_capacity_signals.py

The script auto-detects which directories have files; missing sources are
skipped with a warning, NOT an error. Downstream enrich gracefully handles
absent JSONLs.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

ROOT = Path(__file__).parent.parent
INPUT_DIR = ROOT / "data" / "capacity_inputs"


# ---------- helpers ---------------------------------------------------------

def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        f = float(str(v).replace(",", "").strip())
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def _open_table(path: Path):
    """Yield row dicts from .csv or .xlsx. Handles both formats transparently."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                yield row
    elif suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print(f"  [skip] {path.name}: openpyxl not installed (pip install openpyxl)")
            return
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            # Find header row — first row with no None/empty values, or just take first
            header = None
            data_start = 0
            for idx, row in enumerate(rows[:5]):
                non_empty = sum(1 for v in row if v not in (None, ""))
                if non_empty >= 3:
                    header = [str(v).strip() if v is not None else "" for v in row]
                    data_start = idx + 1
                    break
            if not header:
                continue
            for row in rows[data_start:]:
                yield {h: row[i] if i < len(row) else None for i, h in enumerate(header)}
        wb.close()


def _find_col(row: dict, *patterns: str) -> str | None:
    """Return the value of the first column whose name matches one of the
    given regex patterns (case-insensitive)."""
    for pat in patterns:
        rx = re.compile(pat, re.IGNORECASE)
        for k, v in row.items():
            if k and rx.search(str(k)):
                return v
    return None


def _files_in(subdir: str) -> list[Path]:
    d = INPUT_DIR / subdir
    if not d.exists():
        return []
    return sorted([p for p in d.iterdir() if p.suffix.lower() in (".csv", ".xlsx", ".xls")])


# ---------- PJM queue -------------------------------------------------------

def parse_pjm_queue() -> int:
    files = _files_in("pjm_queue")
    out_path = ROOT / "data" / "pjm_queue.jsonl"
    if not files:
        out_path.write_text("", encoding="utf-8")
        print(f"  [pjm_queue] no files in {INPUT_DIR/'pjm_queue'} (drop xlsx/csv exports there)")
        return 0
    n = 0
    with out_path.open("w", encoding="utf-8") as out:
        for f in files:
            for row in _open_table(f):
                state = (_find_col(row, r"^state$|state\s*$") or "").strip().upper()
                if state and state != "VA":
                    continue
                queue_id = _find_col(row, r"queue\s*(num|id|#)")
                mw = _safe_float(_find_col(row, r"\bmw\b|mfo|capacity"))
                poi = _find_col(row, r"poi|interconnect|substation|point\s*of\s*interconnection")
                status = _find_col(row, r"^status$|withdraw")
                submitted = _find_col(row, r"submitted|date.*entered|received")
                if not (queue_id and (mw or poi)):
                    continue
                rec = {
                    "queue_id": str(queue_id).strip(),
                    "mw": mw,
                    "poi_substation": (str(poi).strip().upper() if poi else ""),
                    "status": (str(status).strip() if status else ""),
                    "submitted": (str(submitted)[:10] if submitted else ""),
                    "source_file": f.name,
                }
                out.write(json.dumps(rec) + "\n")
                n += 1
    print(f"  [pjm_queue] parsed {n} VA-state rows from {len(files)} file(s) -> {out_path}")
    return n


# ---------- EIA Form 860 ----------------------------------------------------

def parse_eia_860() -> int:
    files = _files_in("eia_860")
    out_path = ROOT / "data" / "eia_generators_va.jsonl"
    if not files:
        out_path.write_text("", encoding="utf-8")
        print(f"  [eia_860] no files in {INPUT_DIR/'eia_860'} (drop annual xlsx there)")
        return 0
    n = 0
    with out_path.open("w", encoding="utf-8") as out:
        for f in files:
            for row in _open_table(f):
                state = (_find_col(row, r"^state$") or "").strip().upper()
                if state and state != "VA":
                    continue
                plant = _find_col(row, r"plant\s*name")
                county = _find_col(row, r"^county$")
                mw = _safe_float(_find_col(row, r"nameplate\s*capacity|mw\b"))
                status = _find_col(row, r"^status$|operational")
                op_yr = _safe_float(_find_col(row, r"operating\s*year"))
                ret_yr = _safe_float(_find_col(row, r"retirement\s*year"))
                lat = _safe_float(_find_col(row, r"latitude"))
                lng = _safe_float(_find_col(row, r"longitude"))
                if not (plant and mw):
                    continue
                rec = {
                    "plant_name": str(plant).strip(),
                    "county": (str(county).strip().upper() if county else ""),
                    "nameplate_mw": mw,
                    "status": (str(status).strip() if status else ""),
                    "operating_year": int(op_yr) if op_yr else None,
                    "retirement_year": int(ret_yr) if ret_yr else None,
                    "lat": lat,
                    "lng": lng,
                    "source_file": f.name,
                }
                out.write(json.dumps(rec) + "\n")
                n += 1
    print(f"  [eia_860] parsed {n} VA generator rows from {len(files)} file(s) -> {out_path}")
    return n


# ---------- VADEQ water withdrawals ----------------------------------------

def parse_vadeq_water() -> int:
    files = _files_in("vadeq_water")
    out_path = ROOT / "data" / "va_water_withdrawals.jsonl"
    if not files:
        out_path.write_text("", encoding="utf-8")
        print(f"  [vadeq_water] no files in {INPUT_DIR/'vadeq_water'}")
        return 0
    n = 0
    with out_path.open("w", encoding="utf-8") as out:
        for f in files:
            for row in _open_table(f):
                facility = _find_col(row, r"facility\s*name|user\s*name")
                source_type = _find_col(row, r"source\s*type|water\s*source")
                # Prefer MGD; fall back to MGY (annual gallons in millions) -> divide by 365
                mgd = _safe_float(_find_col(row, r"mgd|gallons.*day"))
                mgy = _safe_float(_find_col(row, r"mgy|gallons.*year|annual.*million"))
                if mgd is None and mgy is not None:
                    mgd = mgy / 365.0
                lat = _safe_float(_find_col(row, r"latitude"))
                lng = _safe_float(_find_col(row, r"longitude"))
                if not (facility and (mgd or mgy)):
                    continue
                rec = {
                    "facility": str(facility).strip(),
                    "source_type": (str(source_type).strip() if source_type else ""),
                    "permitted_mgd": mgd,
                    "lat": lat,
                    "lng": lng,
                    "source_file": f.name,
                }
                out.write(json.dumps(rec) + "\n")
                n += 1
    print(f"  [vadeq_water] parsed {n} VA water withdrawal rows -> {out_path}")
    return n


# ---------- VPDES discharge volumes ----------------------------------------

def parse_vpdes() -> int:
    files = _files_in("vpdes")
    out_path = ROOT / "data" / "va_vpdes_volumes.jsonl"
    if not files:
        out_path.write_text("", encoding="utf-8")
        print(f"  [vpdes] no files in {INPUT_DIR/'vpdes'}")
        return 0
    n = 0
    with out_path.open("w", encoding="utf-8") as out:
        for f in files:
            for row in _open_table(f):
                permit = _find_col(row, r"permit\s*(num|id|#)|npdes")
                facility = _find_col(row, r"facility\s*name")
                permitted = _safe_float(_find_col(row, r"permitted.*flow|design.*flow|max.*flow"))
                avg = _safe_float(_find_col(row, r"avg.*flow|actual.*flow|reported.*flow"))
                lat = _safe_float(_find_col(row, r"latitude"))
                lng = _safe_float(_find_col(row, r"longitude"))
                if not (permit and (permitted or avg)):
                    continue
                rec = {
                    "permit": str(permit).strip(),
                    "facility": (str(facility).strip() if facility else ""),
                    "permitted_mgd": permitted,
                    "avg_mgd": avg,
                    "residual_mgd": (
                        permitted - avg if (permitted is not None and avg is not None) else None
                    ),
                    "lat": lat,
                    "lng": lng,
                    "source_file": f.name,
                }
                out.write(json.dumps(rec) + "\n")
                n += 1
    print(f"  [vpdes] parsed {n} VA VPDES rows -> {out_path}")
    return n


# ---------- FCC broadband fiber --------------------------------------------

def parse_fcc_fiber() -> int:
    files = _files_in("fcc_broadband")
    out_path = ROOT / "data" / "va_fcc_fiber.jsonl"
    if not files:
        out_path.write_text("", encoding="utf-8")
        print(f"  [fcc_fiber] no files in {INPUT_DIR/'fcc_broadband'}")
        return 0
    n = 0
    # FCC schema: location_id, provider_id, brand_name, technology, max_advertised_*,
    # latitude, longitude. Tech codes 50 = FTTP, 71/72 = other terrestrial fixed wireline.
    fiber_codes = {"50", "71", "72"}
    with out_path.open("w", encoding="utf-8") as out:
        for f in files:
            for row in _open_table(f):
                tech = _find_col(row, r"^technology$|tech.*code")
                tech_str = str(tech).strip() if tech is not None else ""
                if tech_str not in fiber_codes:
                    continue
                provider = _find_col(row, r"brand|provider|isp.*name")
                lat = _safe_float(_find_col(row, r"latitude"))
                lng = _safe_float(_find_col(row, r"longitude"))
                if not (provider and lat and lng):
                    continue
                rec = {
                    "provider": str(provider).strip(),
                    "technology": tech_str,
                    "lat": lat,
                    "lng": lng,
                    "source_file": f.name,
                }
                out.write(json.dumps(rec) + "\n")
                n += 1
    print(f"  [fcc_fiber] parsed {n} VA fiber rows -> {out_path}")
    return n


# ---------- main ------------------------------------------------------------

def main() -> int:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    for sub in ("pjm_queue", "eia_860", "vadeq_water", "vpdes", "fcc_broadband"):
        (INPUT_DIR / sub).mkdir(parents=True, exist_ok=True)

    print("=== capacity-signal ingest ===")
    parse_pjm_queue()
    parse_eia_860()
    parse_vadeq_water()
    parse_vpdes()
    parse_fcc_fiber()
    print()
    print(f"Drop new bulk files under {INPUT_DIR}/<source>/ and re-run.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
